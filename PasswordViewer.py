from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button

from pyperclip import copy
from WarningPopup import WarningPopup
from kivy.uix.popup import Popup

from openpyxl import load_workbook


class PasswordViewer(Popup):
    def __init__(self, path, **kwargs):
        super(PasswordViewer, self).__init__(**kwargs)

        self.bl = BoxLayout(orientation="vertical", spacing=1)
        self.__path = path
        self.password_book = None

        try:
            self.password_book = load_workbook(self.__path)
            password = self.password_book.active

            for row_id, cell in enumerate(password.iter_rows(min_row=0, max_row=1000,
                                               min_col=1, max_col=4, values_only=True)):

                    if cell.count(None) > 0:
                        break

                    password_bl = BoxLayout(size_hint=(1, 0.3))

                    password_bl.add_widget(Label(text=str(row_id), font_size=20))
                    password_bl.add_widget(Label(text=cell[0], font_size=20))

                    password_bl.add_widget(Button(text="Copy password", on_press=self.copy_password,
                                                  font_size=15, background_color=[0.4, 0.9, 0, 1]))

                    password_bl.add_widget(Button(text="Delete password", on_press=self.delete_password,
                                                  font_size=15, background_color=[0.9, 0.1, 0, 1]))

                    password_bl.add_widget(Label(text=cell[1], font_size=20))
                    password_bl.add_widget(Label(text=cell[2], font_size=20))

                    password_bl.add_widget(Label(text=cell[3], font_size=20))

                    self.bl.add_widget(password_bl)

        except Exception as exc:
            print(exc)

        self.title_align = "center"
        self.title = "Passwords"

        self.content = self.bl
        self.bl.add_widget(Button(text="Close", on_press=self.dismiss, size_hint=(1, 0.2)))

    @staticmethod
    def copy_password(instance):
        password_widget = instance.parent.children[1]

        try:
            if password_widget.text != "" or None:
                WarningPopup("Password copied")
                copy(password_widget.text)

            else:
                WarningPopup("Wrong password, can not copy!")

        except Exception as exc:
            if exc is not None or "":
                WarningPopup("Error with password")
                print(exc)

    def delete_password(self, instance):
        try:
            row_id = instance.parent.children[6].text
            self.password_book.active.delete_rows(int(row_id))

            self.password_book.save(self.__path)
            self.bl.remove_widget(instance.parent)

        except Exception as exc:
            print(exc)
            WarningPopup("Error, cant delete")
